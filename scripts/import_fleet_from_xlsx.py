from __future__ import annotations

import argparse
import json
import os
import re
import sys
import xmlrpc.client
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_URL = "http://127.0.0.1:8071"
DEFAULT_DB = "odoo19_admin"
DEFAULT_LOGIN = "admin"
DEFAULT_PASSWORD = "admin"

BRAND_RULES: list[tuple[str, str]] = [
    ("bongo", "Kia"),
    ("kamaz", "KAMAZ"),
    ("dong feng", "Dongfeng"),
    ("dongfeng", "Dongfeng"),
    ("hyundai", "Hyundai"),
    ("beiben", "Beiben"),
    ("north benz", "North Benz"),
    ("хово", "North Benz"),
    ("chenli", "Chenli"),
    ("sinotruck", "Sinotruck"),
    ("foton", "Foton"),
    ("faw", "FAW"),
]


@dataclass
class FleetVehicleRow:
    license_plate: str
    model_name: str
    name: str


class OdooClient:
    def __init__(self, url: str, db: str, login: str, password: str) -> None:
        self.url = url.rstrip("/")
        self.db = db
        self.login = login
        self.password = password
        self.common = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/common")
        self.uid = self.common.authenticate(self.db, self.login, self.password, {})
        if not self.uid:
            raise RuntimeError("Odoo нэвтрэлт амжилтгүй боллоо.")
        self.models = xmlrpc.client.ServerProxy(f"{self.url}/xmlrpc/2/object")

    def execute(self, model: str, method: str, args: list | None = None, kwargs: dict | None = None):
        return self.models.execute_kw(
            self.db,
            self.uid,
            self.password,
            model,
            method,
            args or [],
            kwargs or {},
        )


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("_x000d_", " ")
    text = text.replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def infer_brand_name(model_name: str) -> str:
    lowered = model_name.casefold()
    for needle, brand_name in BRAND_RULES:
        if needle in lowered:
            return brand_name
    first_token = clean_text(model_name).split(" ", 1)[0]
    return first_token or "Unknown"


def load_workbook_rows(xlsx_path: Path) -> tuple[list[FleetVehicleRow], list[str]]:
    workbook = load_workbook(xlsx_path, data_only=True)

    if "vehicles_import" not in workbook.sheetnames:
        raise RuntimeError("Excel файлд 'vehicles_import' sheet алга.")

    vehicle_sheet = workbook["vehicles_import"]
    vehicle_rows = list(vehicle_sheet.iter_rows(values_only=True))
    if not vehicle_rows:
        raise RuntimeError("vehicles_import sheet хоосон байна.")

    header = [clean_text(cell) for cell in vehicle_rows[0]]
    expected = ["license_plate", "model_id", "name"]
    if header[:3] != expected:
        raise RuntimeError(f"Header таарахгүй байна: {header!r}")

    vehicles: list[FleetVehicleRow] = []
    seen_license_plates: set[str] = set()

    for row_number, row in enumerate(vehicle_rows[1:], start=2):
        license_plate = clean_text(row[0] if len(row) > 0 else "")
        model_name = clean_text(row[1] if len(row) > 1 else "")
        name = clean_text(row[2] if len(row) > 2 else "")

        if not license_plate and not model_name and not name:
            continue
        if not license_plate or not model_name:
            raise RuntimeError(f"{row_number}-р мөр дээр улсын дугаар эсвэл model дутуу байна.")
        if license_plate in seen_license_plates:
            raise RuntimeError(f"Давхардсан улсын дугаар илэрлээ: {license_plate}")

        seen_license_plates.add(license_plate)
        vehicles.append(
            FleetVehicleRow(
                license_plate=license_plate,
                model_name=model_name,
                name=name or f"{model_name} {license_plate}",
            )
        )

    model_names = sorted({vehicle.model_name for vehicle in vehicles}, key=str.casefold)

    if "models" in workbook.sheetnames:
        model_sheet = workbook["models"]
        supplied_models = [
            clean_text(row[0])
            for row in model_sheet.iter_rows(min_row=2, values_only=True)
            if row and clean_text(row[0])
        ]
        if supplied_models:
            supplied_set = {clean_text(name) for name in supplied_models}
            missing_models = sorted(set(model_names) - supplied_set, key=str.casefold)
            if missing_models:
                raise RuntimeError(
                    "models sheet дотор vehicles_import sheet-ийн model дутуу байна: "
                    + ", ".join(missing_models)
                )
            model_names = sorted(supplied_set, key=str.casefold)

    return vehicles, model_names


def backup_records(client: OdooClient, backup_root: Path) -> Path:
    backup_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_root / f"fleet_backup_{timestamp}.json"

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "vehicles": client.execute(
            "fleet.vehicle",
            "search_read",
            [[]],
            {
                "fields": [
                    "id",
                    "name",
                    "license_plate",
                    "model_id",
                    "state_id",
                    "company_id",
                    "driver_id",
                    "future_driver_id",
                ],
                "order": "id",
            },
        ),
        "models": client.execute(
            "fleet.vehicle.model",
            "search_read",
            [[]],
            {
                "fields": ["id", "name", "brand_id", "vehicle_type"],
                "order": "id",
            },
        ),
    }

    backup_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return backup_path


def ensure_brands(client: OdooClient, brand_names: Iterable[str]) -> dict[str, int]:
    desired = sorted({clean_text(name) for name in brand_names if clean_text(name)}, key=str.casefold)
    existing = client.execute(
        "fleet.vehicle.model.brand",
        "search_read",
        [[["name", "in", desired]]],
        {"fields": ["id", "name"], "limit": len(desired) or 1},
    )
    brand_map = {record["name"]: record["id"] for record in existing}

    for brand_name in desired:
        if brand_name in brand_map:
            continue
        brand_id = client.execute(
            "fleet.vehicle.model.brand",
            "create",
            [{"name": brand_name}],
        )
        brand_map[brand_name] = brand_id

    return brand_map


def upsert_models(
    client: OdooClient,
    model_names: list[str],
    brand_map: dict[str, int],
) -> dict[str, int]:
    existing_models = client.execute(
        "fleet.vehicle.model",
        "search_read",
        [[]],
        {"fields": ["id", "name"], "order": "id"},
    )
    existing_by_name = {record["name"]: record["id"] for record in existing_models}

    model_map: dict[str, int] = {}
    for model_name in model_names:
        existing_id = existing_by_name.get(model_name)
        values = {
            "name": model_name,
            "brand_id": brand_map[infer_brand_name(model_name)],
            "vehicle_type": "car",
        }
        if existing_id:
            client.execute("fleet.vehicle.model", "write", [[existing_id], values])
            model_map[model_name] = existing_id
            continue
        model_id = client.execute("fleet.vehicle.model", "create", [values])
        model_map[model_name] = model_id

    return model_map


def prune_stale_models(client: OdooClient, model_names: list[str]) -> None:
    desired_model_names = set(model_names)
    existing_models = client.execute(
        "fleet.vehicle.model",
        "search_read",
        [[]],
        {"fields": ["id", "name"], "order": "id"},
    )
    removable_ids = [
        record["id"]
        for record in existing_models
        if record["name"] not in desired_model_names
    ]
    if removable_ids:
        client.execute("fleet.vehicle.model", "unlink", [removable_ids])


def resolve_company_id(client: OdooClient) -> int:
    companies = client.execute(
        "res.company",
        "search_read",
        [[]],
        {"fields": ["id", "name"], "limit": 1, "order": "id"},
    )
    if not companies:
        raise RuntimeError("res.company дотор компани олдсонгүй.")
    return companies[0]["id"]


def clear_vehicles(client: OdooClient) -> None:
    vehicle_ids = client.execute("fleet.vehicle", "search", [[]])
    if vehicle_ids:
        client.execute("fleet.vehicle", "unlink", [vehicle_ids])


def create_vehicles(
    client: OdooClient,
    vehicles: list[FleetVehicleRow],
    model_map: dict[str, int],
    company_id: int,
) -> list[int]:
    created_ids: list[int] = []
    for vehicle in vehicles:
        created_id = client.execute(
            "fleet.vehicle",
            "create",
            [
                {
                    "name": vehicle.name,
                    "license_plate": vehicle.license_plate,
                    "model_id": model_map[vehicle.model_name],
                    "company_id": company_id,
                }
            ],
        )
        created_ids.append(created_id)
    return created_ids


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Replace Odoo fleet vehicles from an XLSX workbook.")
    parser.add_argument("--xlsx", required=True, help="Path to the Excel workbook.")
    parser.add_argument("--url", default=os.environ.get("ODOO_URL", DEFAULT_URL))
    parser.add_argument("--db", default=os.environ.get("ODOO_DB", DEFAULT_DB))
    parser.add_argument("--login", default=os.environ.get("ODOO_LOGIN", DEFAULT_LOGIN))
    parser.add_argument("--password", default=os.environ.get("ODOO_PASSWORD", DEFAULT_PASSWORD))
    parser.add_argument(
        "--backup-dir",
        default=str(Path(__file__).resolve().parent / "backups"),
        help="Directory for JSON backups before replacement.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise RuntimeError(f"Excel файл олдсонгүй: {xlsx_path}")

    client = OdooClient(args.url, args.db, args.login, args.password)
    vehicles, model_names = load_workbook_rows(xlsx_path)

    backup_path = backup_records(client, Path(args.backup_dir).resolve())
    brand_names = sorted({infer_brand_name(name) for name in model_names}, key=str.casefold)
    brand_map = ensure_brands(client, brand_names)
    model_map = upsert_models(client, model_names, brand_map)
    clear_vehicles(client)
    prune_stale_models(client, model_names)
    company_id = resolve_company_id(client)
    created_vehicle_ids = create_vehicles(client, vehicles, model_map, company_id)

    summary = {
        "backup_path": str(backup_path),
        "vehicle_count": len(created_vehicle_ids),
        "model_count": len(model_map),
        "brands": sorted(brand_map.keys(), key=str.casefold),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # pragma: no cover - CLI guard
        print(f"ERROR: {error}", file=sys.stderr)
        raise
